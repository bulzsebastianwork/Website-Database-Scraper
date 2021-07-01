import os
import requests
import threading
from bs4 import BeautifulSoup
from openpyxl import load_workbook

products_info_list = []

def main():
    products_text_file = open("Log_File_Products.txt", "r")
    lines_products_file = products_text_file.readlines()

    list_of_links = []

    for line in lines_products_file:
        list_of_links.append(line.split("link: ")[-1].replace("\n", ""))


    if len(list_of_links) == 1:
        category = ""

        get_all_product_info(list_of_links[0], category)
    else:
        # If there are more products multi-thread the instances
        i = 0
        while i < len(list_of_links):
            max_threads_to_spawn = 15  # Number of maximum instances to spawn
            thread_list = list()

            # Start test
            for j in range(max_threads_to_spawn):
                if i < len(list_of_links):
                    category = ""

                    t = threading.Thread(name='Test {}'.format(i), target=get_all_product_info, args=(list_of_links[i], category))
                    t.start()
                    thread_list.append(t)

                    i += 1
                else:
                    break

            # Wait for all threads to complete
            for thread in thread_list:
                thread.join()


    wb_export = load_workbook("Export_Database.xlsx")

    ws_export = wb_export["Export_Sheet"]

    counter_rows = ws_export.max_row + 1

    start_videos_column = 1
    while not ("Youtube" in ws_export.cell(1, start_videos_column).value):
        start_videos_column += 1

    application_column = start_videos_column
    while ws_export.cell(1, application_column).value != "Application":
        application_column += 1

    for i in range(len(products_info_list)):
        ws_export.cell(counter_rows + i, 1).value = products_info_list[i].sku
        ws_export.cell(counter_rows + i, 2).value = products_info_list[i].title
        ws_export.cell(counter_rows + i, 3).value = products_info_list[i].category

        if len(products_info_list[i].images_list) > 0:
            for image in products_info_list[i].images_list:
                if os.path.getsize("Images/" + image) > 1000:
                    ws_export.cell(counter_rows + i, image_number(image) + 4).value = image
                else:
                    try:
                        print.write("Sku: " + products_info_list[i].sku + ",image " + str(image_number(image)) + ", was deleted because it was corrupted\n")
                        os.remove("Images/" + image)
                    except:
                        print(image + " cannot be deleted, no privileges on the app.")
        if len(products_info_list[i].video_links_list) > 0:
            for video in products_info_list[i].video_links_list:
                ws_export.cell(counter_rows + i, products_info_list[i].video_links_list.index(video) + start_videos_column).value = video

        ws_export.cell(counter_rows + i, application_column).value = products_info_list[i].application

        if "IN STOCK" in products_info_list[i].stock.upper():
            ws_export.cell(counter_rows + i, application_column + 1).value = "In Stock"
        elif "OUT OF STOCK" in products_info_list[i].stock.upper():
            ws_export.cell(counter_rows + i, application_column + 1).value = "Out of Stock"
        else:
            ws_export.cell(counter_rows + i, application_column + 1).value = "HTML Element not found"


    wb_export.save("Export_Database_w_rest.xlsx")

    print("Done!")


def get_all_product_info(link, category):
    try:
        html_from_link = requests.get(link)
        soup = BeautifulSoup(html_from_link.text, 'html.parser')

        sku = soup.find("td", {"class": "product-details-value"}).text
        title = soup.find("h1", {"class": "productView-title"}).text

        list_of_images_carousel = soup.find_all("li", {"class":"productView-imageCarousel-main-item"})

        images_links_list = []
        if not(list_of_images_carousel is None):
            for image_slider in list_of_images_carousel:
                images_links_list.append(image_slider.find("a").get("href"))

        images_list = []

        i = 0
        while i < len(images_links_list):
            max_threads_to_spawn = 7  # Number of maximum browsers to spawn
            thread_list = list()

            # Start test
            for j in range(max_threads_to_spawn):
                if i < len(images_links_list):
                    t = threading.Thread(name='Image Download {}'.format(i), target=download_and_rename_image,
                                         args=(sku, images_links_list[i], i + 1, images_list))
                    t.start()
                    thread_list.append(t)

                    i += 1
                else:
                    break

            # Wait for all threads to complete
            for thread in thread_list:
                thread.join()

        images_list.sort()

        youtube_links_list = []

        list_of_videos_caroussel = soup.find_all("a", {"class":"video"})

        if not(list_of_videos_caroussel is None):
            for video in list_of_videos_caroussel:
                youtube_links_list.append("https://www.youtube.com/watch?v=" + video.get("data-video-id"))

        description_div_element = soup.text.split("Vehicle Fitment:")[-1].split("\n\n\n")[0]

        try:
            if description_div_element[0:1] == "\n":
                description_div_element = description_div_element[2:]
            elif description_div_element[0] == " ":
                description_div_element = description_div_element[1:]
        except:
            description_div_element = ""

        try:
            stock_element = soup.find("p", {"class":"alertBox-message"})
            stock_text = stock_element.find("span").text
        except:
            stock_text = ""

        products_info_list.append(Product(sku, title, category, images_list, youtube_links_list, description_div_element, stock_text))
    except:
        print("Product wasn't gathered from link: " + link + "\n")


def download_and_rename_image(sku, image_link, image_number, images_list):
    image_extension = image_link.split(".")[-1].split("?")[0]

    if image_number == 1:
        image_filename = sku + "." + image_extension
    else:
        image_filename = sku + "-" + str(image_number - 1) + "." + image_extension

    try:
        r = requests.get(image_link, allow_redirects=True, timeout=5)

        open("Images\\" + image_filename, 'wb').write(r.content)

        images_list.append(image_filename)
    except:
        print("Sku: " + sku + ",image " + str(image_number) + ",link: " + image_link + " couldn't be downloaded\n")

def image_number(image_string):
    if image_string.split(".")[0].split("-")[-1].isnumeric():
        if int(image_string.split(".")[0].split("-")[-1]) > 20:
            return 0
        else:
            return int(image_string.split(".")[0].split("-")[-1])
    else:
        return 0

class Product():
    def __init__(self, sku, title, category, images_list, video_links_list, application, stock):
        self.sku = sku
        self.title = title
        self.category = category
        self.images_list = images_list
        self.video_links_list = video_links_list
        self.application = application
        self.stock = stock

main()