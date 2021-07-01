import os
import requests
import threading
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


products_info_list = []

def main():

    chrome_options = Options()

    chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(options=chrome_options)
    start_url = "https://specdtuning.com/"
    driver.get(start_url)

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'navPage-subMenu-item')))
    all_links_parents = driver.find_elements_by_class_name("navPage-subMenu-action")

    all_links_childs = driver.find_elements_by_class_name("navPage-childList-action")

    all_links_set = set()
    for link in all_links_parents:
        if not (link.get_attribute("href") is None):
            all_links_set.add(link.get_attribute("href"))

    for link in all_links_childs:
        if not (link.get_attribute("href") is None):
            all_links_set.add(link.get_attribute("href"))

    driver.close()

    all_links_list = list(all_links_set)

    all_links_list.sort()

    log_file = open("Log_File_Products.txt", "w+")
    log_file_images = open("Log_File_Images.txt", "w+")

    for link in all_links_list:
        print("Progress "  + "{:.2f}".format(float(all_links_list.index(link) / len(all_links_list) * 100)) + "% Completed")
        print("Getting products from link: " + link)
        get_all_product_links_from_page(link, log_file, log_file_images)

    wb = Workbook()
    ws = wb.active
    ws.title = "Export_Sheet"

    ws.cell(1, 1).value = "Sku"
    ws.cell(1, 2).value = "Title"
    ws.cell(1, 3).value = "Category"


    max_images = 0
    max_youtube_links = 0
    # Finding the max number of images and youtube links for one product
    for product in products_info_list:
        if len(product.images_list) > max_images:
            max_images = len(product.images_list)

        if len(product.video_links_list) > max_youtube_links:
            max_youtube_links = len(product.video_links_list)

    ws.cell(1, 4).value = "Main Image"

    # Adding the images and videos headers to the worksheet
    for i in range(2, max_images + 1):
        ws.cell(1, i + 3).value = "Additional Image " + str(i - 1)

    start_videos_column = max_images + 3
    for i in range(1, max_youtube_links + 1):
        ws.cell(1, start_videos_column + i).value = "Youtube " + str(i)

    ws.cell(1, start_videos_column + max_youtube_links + 1).value = "Application"
    ws.cell(1, start_videos_column + max_youtube_links + 2).value = "Availability"

    for i in range(len(products_info_list)):
        ws.cell(i + 2, 1).value = products_info_list[i].sku
        ws.cell(i + 2, 2).value = products_info_list[i].title
        ws.cell(i + 2, 3).value = products_info_list[i].category

        if len(products_info_list[i].images_list) > 0:
            for image in products_info_list[i].images_list:
                if os.path.getsize("Images/" + image) > 1000:
                    ws.cell(i + 2, image_number(image) + 4).value = image
                else:
                    try:
                        log_file_images.write("Sku: " + products_info_list[i].sku + ",image " + str(image_number(image)) + ", was deleted because it was corrupted\n")
                        os.remove("Images/" + image)
                    except:
                        print(image + " cannot be deleted, no privileges on the app.")
        if len(products_info_list[i].video_links_list) > 0:
            for video in products_info_list[i].video_links_list:
                ws.cell(i + 2, products_info_list[i].video_links_list.index(video) + start_videos_column + 1).value = video

        if "IN STOCK" in products_info_list[i].stock.upper():
            ws.cell(i + 2, start_videos_column + max_youtube_links + 2).value = "In Stock"
        elif "OUT OF STOCK" in products_info_list[i].stock.upper():
            ws.cell(i + 2, start_videos_column + max_youtube_links + 2).value = "Out of Stock"
        else:
            ws.cell(i + 2, start_videos_column + max_youtube_links + 2).value = "HTML Element not found"

        ws.cell(i + 2, start_videos_column + max_youtube_links + 1).value = products_info_list[i].application

    log_file.close()
    log_file_images.close()

    wb.save("Export_Database.xlsx")


def get_all_product_links_from_page(link, log_file, log_file_images):
    chrome_options = Options()
    # chrome_options.add_argument("--headless")

    driver = webdriver.Chrome(options=chrome_options)
    driver.get(link)


    nothing_found = driver.find_elements_by_link_text("Nothing found")
    if not(nothing_found is None):
        try:
            main_results_element = driver.find_element_by_class_name("snize-main-panel-collection-results-found")
            results = int(main_results_element.find_element_by_tag_name("b").text)
        except:
            results = 0

        link_changed = False
        if results > 20:
            if results > 200:
                link_changed = True
                link = link + "/?tab=products&page=10"
            else:
                link_changed = True
                pages = results // 20

                if pages % 20 != 0:
                    pages += 1
                link = link + "/?tab=products&page=" + str(pages)


        if link_changed == True:
            driver.get(link)

        show_more_button = driver.find_elements_by_link_text("Show more")
        while len(show_more_button) > 0:
            show_more_button[0].click()

            try:
                WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.LINK_TEXT, 'Show more')))
            except:
                break

            show_more_button = driver.find_elements_by_link_text("Show more")

        all_visible_product_links = driver.find_elements_by_class_name("snize-view-link")
        list_of_links = []

        for link_element in all_visible_product_links:
            list_of_links.append(link_element.get_attribute("href"))

        driver.close()

        if link.split("/")[-2] != "":
            category = link.split("/")[-2]
        else:
            if len(link.split("/")) > 2:
                category = link.split("/")[-3]
            else:
                category = ""

        if len(list_of_links) == 1:
            get_all_product_info(list_of_links[0], category, log_file, log_file_images)
        else:
            # If there are more products multi-thread the instances
            i = 0
            while i < len(list_of_links):
                max_threads_to_spawn = 15  # Number of maximum instances to spawn
                thread_list = list()

                # Start test
                for j in range(max_threads_to_spawn):
                    if i < len(list_of_links):
                        t = threading.Thread(name='Test {}'.format(i), target=get_all_product_info, args=(list_of_links[i], category, log_file, log_file_images))
                        t.start()
                        thread_list.append(t)

                        i += 1
                    else:
                        break

                # Wait for all threads to complete
                for thread in thread_list:
                    thread.join()
    else:
        driver.close()



def get_all_product_info(link, category, log_file, log_file_images):
    try:
        html_from_link = requests.get(link)
        soup = BeautifulSoup(html_from_link.text, 'html.parser')

        sku = soup.find("td", {"class": "product-details-value"}).text
        title = soup.find("h1", {"class": "productView-title"}).text

        list_of_images_carousel = soup.find_all("li", {"class":"productView-imageCarousel-main-item"})

        images_links_list = []
        if not(list_of_images_carousel is None):
            for image_slider in list_of_images_carousel:
                images_links_list.append(image_slider.find("a").get("href"))

        images_list = []

        i = 0
        while i < len(images_links_list):
            max_threads_to_spawn = 7  # Number of maximum browsers to spawn
            thread_list = list()

            # Start test
            for j in range(max_threads_to_spawn):
                if i < len(images_links_list):
                    t = threading.Thread(name='Image Download {}'.format(i), target=download_and_rename_image,
                                         args=(sku, images_links_list[i], i + 1, images_list, log_file_images))
                    t.start()
                    thread_list.append(t)

                    i += 1
                else:
                    break

            # Wait for all threads to complete
            for thread in thread_list:
                thread.join()

        images_list.sort()

        youtube_links_list = []

        list_of_videos_caroussel = soup.find_all("a", {"class":"video"})

        if not(list_of_videos_caroussel is None):
            for video in list_of_videos_caroussel:
                youtube_links_list.append("https://www.youtube.com/watch?v=" + video.get("data-video-id"))

        description_div_element = soup.text.split("Vehicle Fitment:")[-1].split("\n\n\n")[0]

        try:
            if description_div_element[0:1] == "\n":
                description_div_element = description_div_element[2:]
            elif description_div_element[0] == " ":
                description_div_element = description_div_element[1:]
        except:
            description_div_element = ""

        try:
            stock_element = soup.find("p", {"class":"alertBox-message"})
            stock_text = stock_element.find("span").text
        except:
            stock_text = ""

        products_info_list.append(Product(sku, title, category, images_list, youtube_links_list, description_div_element, stock_text))
    except:
        log_file.write("Product wasn't gathered from link: " + link + "\n")


def download_and_rename_image(sku, image_link, image_number, images_list, log_file_images):
    image_extension = image_link.split(".")[-1].split("?")[0]

    if image_number == 1:
        image_filename = sku + "." + image_extension
    else:
        image_filename = sku + "-" + str(image_number - 1) + "." + image_extension

    try:
        r = requests.get(image_link, allow_redirects=True, timeout=5)

        open("Images\\" + image_filename, 'wb').write(r.content)

        images_list.append(image_filename)
    except:
        log_file_images.write("Sku: " + sku + ",image " + str(image_number) + ",link: " + image_link + " couldn't be downloaded\n")

def image_number(image_string):
    if image_string.split(".")[0].split("-")[-1].isnumeric():
        if int(image_string.split(".")[0].split("-")[-1]) > 20:
            return 0
        else:
            return int(image_string.split(".")[0].split("-")[-1])
    else:
        return 0

class Product():
    def __init__(self, sku, title, category, images_list, video_links_list, application, stock):
        self.sku = sku
        self.title = title
        self.category = category
        self.images_list = images_list
        self.video_links_list = video_links_list
        self.application = application
        self.stock = stock
main()