import os
import requests
import threading
from openpyxl import load_workbook

def main():
    images_text_file = open("Log_File_Images.txt", "r")

    list_of_lines_images_file = images_text_file.readlines()

    list_of_image_objects = []

    for line in list_of_lines_images_file:
        split_comma = line.split(",")

        if "link" in line:
            list_of_image_objects.append(sku_with_image(split_comma[0].split(" ")[-1], int(split_comma[1].split(" ")[-1]), split_comma[2].split(" ")[1]))

    wb_export = load_workbook("Export_Database.xlsx")
    ws_export = wb_export["Export_Sheet"]

    dict_of_skus = {}

    counter_rows = 2
    while counter_rows <= ws_export.max_row:
        dict_of_skus[ws_export.cell(counter_rows, 1).value] = counter_rows

        counter_rows += 1

    for object in list_of_image_objects:
        ws_export.cell(dict_of_skus[object.sku], 4 + object.image_number).value = download_and_rename_image(object.sku, object.link, object.image_number + 1)

    wb_export.save("Export_Database_w_images.xlsx")

def download_and_rename_image(sku, image_link, image_number):
    image_extension = image_link.split(".")[-1].split("?")[0]

    if image_number == 1:
        image_filename = sku + "." + image_extension
    else:
        image_filename = sku + "-" + str(image_number - 1) + "." + image_extension

    try:
        r = requests.get(image_link, allow_redirects=True, timeout=5)

        open("Images\\" + image_filename, 'wb').write(r.content)
    except:
        print("Image " + image_filename + " couldn't be downloaded.")

    return image_filename

class sku_with_image():
    def __init__(self, sku, image_number, link):
        self.sku = sku
        self.image_number = image_number
        self.link = link


main()